from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from news_scraper import excel_exporter
from news_scraper.quality import process_news_quality
from news_scraper.scrapers.ministry.health import cdc


FIXTURES = Path(__file__).parent / "fixtures"


def test_fixture_html_to_quality_checked_excel(monkeypatch, tmp_path):
    html = (FIXTURES / "cdc_list.html").read_text(encoding="utf-8")
    monkeypatch.setattr(cdc, "fetch_html", lambda *args, **kwargs: html)
    monkeypatch.setattr(cdc, "get_cached_week_range", lambda: (date(2026, 6, 8), date(2026, 6, 14)))
    monkeypatch.setattr(excel_exporter, "get_cached_week_range", lambda: (date(2026, 6, 8), date(2026, 6, 14)))

    parsed = cdc.scrape_cdc_this_week()
    cleaned, summary = process_news_quality(parsed, ["疾管署"])
    output = excel_exporter.export_to_excel(cleaned, tmp_path)
    workbook = load_workbook(output, data_only=False)

    assert summary["output_count"] == 1
    assert workbook.sheetnames
    assert workbook["全部新聞"]["D2"].value == "測試防疫新聞"
    assert workbook["全部新聞"]["E2"].hyperlink.target.endswith("example?typeid=9")

