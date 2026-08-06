import json
from datetime import date
from pathlib import Path

try:
    import tomllib
except ModuleNotFoundError:  # pragma: no cover - Python 3.10 compatibility
    import tomli as tomllib

from openpyxl import load_workbook

from news_scraper.color_utils import contrast_ratio, readable_text_color
from news_scraper.errors import ParserContractError
from news_scraper.excel_exporter import export_to_excel
from news_scraper.gui import load_settings
from news_scraper.health_dashboard import build_health_dashboard_model
from news_scraper.monitoring import RunContext, use_run_context
from news_scraper.relevance import build_default_relevance_profile
from news_scraper.source_catalog import SourceRoute, execute_source_routes
from news_scraper.update_checker import check_latest_release, version_key
from news_scraper.version import CURRENT_VERSION


def test_source_route_falls_back_and_records_evidence():
    routes = (
        SourceRoute("primary", "https://primary.example.test/news", priority=1),
        SourceRoute("rss", "https://backup.example.test/rss", kind="rss", priority=2),
    )
    context = RunContext()

    def handler(route):
        if route.id == "primary":
            raise ParserContractError(
                "missing list",
                url=route.url,
                content_type="text/html",
                response_bytes=120,
                selector=".news",
            )
        return [{"source": "測試來源"}]

    with use_run_context(context):
        result = execute_source_routes("測試來源", routes, handler)

    assert result == [{"source": "測試來源"}]
    assert [item.status for item in context.route_attempts] == ["failed", "success"]
    assert context.route_attempts[0].failure_class == "parser_regression"
    assert context.final_routes["測試來源"]["route_id"] == "rss"
    assert context.final_routes["測試來源"]["used_fallback"] is True


def test_custom_topic_color_only_overrides_topic_cell(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 7, 27), date(2026, 8, 2)),
    )
    profile = build_default_relevance_profile()
    profile.topics[0].display_color = "#123456"
    output = export_to_excel(
        [
            {
                "source": "行政院",
                "date": "2026-07-29",
                "department": "行政院",
                "title": "全民智慧生活圈正式啟動",
                "link": "https://example.test/news",
            }
        ],
        output_dir=tmp_path,
        relevance_profile=profile,
    )
    sheet = load_workbook(output)["全部新聞"]

    assert sheet["A2"].fill.fgColor.rgb[-6:] == "FFFF00"
    assert sheet["H2"].fill.fgColor.rgb[-6:] == "123456"
    assert sheet["H2"].font.color.rgb[-6:] == "FFFFFF"
    reference_headers = [cell.value for cell in load_workbook(output)["主題規則對照"][1]]
    assert "顯示顏色" in reference_headers


def test_topic_text_color_always_meets_wcag_contrast():
    for background in ("#000000", "#FFFFFF", "#777777", "#FFFF00", "#123456"):
        foreground, ratio = readable_text_color(background)
        assert contrast_ratio(background, foreground) == ratio
        assert ratio >= 4.5


def test_gui_settings_schema_two_migrates_recent_colors(tmp_path):
    path = tmp_path / "settings.json"
    path.write_text(
        json.dumps(
            {
                "schema_version": 2,
                "sources": ["行政院"],
                "output_dir": str(tmp_path),
                "recent_topic_colors": ["#123456", "#123456", "bad", "#ABCDEF"],
            }
        ),
        encoding="utf-8",
    )

    settings = load_settings(path, ["行政院", "財政部"])

    assert settings.schema_version == 3
    assert settings.recent_topic_colors == ["#123456", "#ABCDEF"]


def test_health_dashboard_uses_final_source_results():
    reports = [
        {
            "status": "partial_failure",
            "finished_at": "2026-07-30T10:00:00+08:00",
            "selected_sources": ["行政院", "財政部"],
            "failed_sources": ["財政部"],
            "insecure_ssl_hosts": [],
            "source_health": {"healthy_count": 1, "failed_count": 1, "fallback_source_count": 1},
            "source_diagnostics": [
                {
                    "source": "行政院",
                    "status": "success",
                    "item_count": 2,
                    "elapsed_seconds": 1.0,
                    "final_route": {"route_id": "primary"},
                },
                {
                    "source": "財政部",
                    "status": "failed",
                    "item_count": 0,
                    "elapsed_seconds": 3.0,
                    "failure_class": "source_outage",
                    "final_route": {"route_id": "alternate-2", "used_fallback": True},
                },
            ],
        }
    ]

    model = build_health_dashboard_model(reports, ["行政院", "財政部"])
    rows = {row["source"]: row for row in model["source_rows"]}

    assert model["healthy_count"] == 1
    assert model["failed_count"] == 1
    assert rows["行政院"]["success_rate"] == 1
    assert rows["財政部"]["success_rate"] == 0
    assert rows["財政部"]["last_failure_class"] == "source_outage"
    assert rows["財政部"]["failure_count"] == 1
    assert rows["財政部"]["latest_elapsed_seconds"] == 3


def test_manual_update_checker_compares_release_versions(monkeypatch):
    requested_urls = []

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            return False

        def read(self):
            return json.dumps(
                {
                    "tag_name": "v2.0.1",
                    "published_at": "2026-08-01T00:00:00Z",
                    "html_url": "https://github.com/example/releases/v2.0.1",
                    "draft": False,
                    "prerelease": False,
                }
            ).encode()

    def fake_urlopen(request, timeout):
        requested_urls.append(request.full_url)
        return FakeResponse()

    monkeypatch.setattr("news_scraper.update_checker.urlopen", fake_urlopen)

    result = check_latest_release()

    assert version_key("v1.6.0") == (1, 6, 0)
    assert result.latest_version == "2.0.1"
    assert result.update_available is True
    assert result.published_at == "2026-08-01"
    assert requested_urls == [
        "https://api.github.com/repos/ECJura2000/taiwan-government-news-scraper/releases/latest"
    ]


def test_runtime_version_matches_project_version():
    project = tomllib.loads(
        (Path(__file__).resolve().parents[1] / "pyproject.toml").read_text(
            encoding="utf-8"
        )
    )

    assert CURRENT_VERSION == project["project"]["version"]
