from datetime import date

from openpyxl import load_workbook

from news_scraper.excel_exporter import export_to_excel


def test_export_groups_affiliated_agencies_under_parent_source(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 4, 20), date(2026, 4, 26)),
    )

    output_path = export_to_excel(
        [
            {
                "source": "消防署",
                "date": "2026-04-23",
                "department": "災害管理組",
                "title": "比較新的消防新聞",
                "link": "https://example.com/fire-new",
            },
            {
                "source": "外交部",
                "date": "2026-04-21",
                "department": "外交部",
                "title": "外交新聞",
                "link": "https://example.com/foreign",
            },
            {
                "source": "內政部",
                "date": "2026-04-22",
                "department": "國土管理署",
                "title": "內政部本部新聞",
                "link": "https://example.com/interior",
            },
            {
                "source": "消防署",
                "date": "2026-04-21",
                "department": "災害管理組",
                "title": "比較舊的消防新聞",
                "link": "https://example.com/fire-old",
            },
        ],
        output_dir=tmp_path,
    )

    workbook = load_workbook(output_path, rich_text=True)
    rows = list(workbook["全部新聞"].iter_rows(min_row=2, values_only=True))

    assert [row[0] for row in rows] == ["內政部", "內政部", "內政部", "外交部"]
    assert [str(row[1]) for row in rows[:3]] == [
        "2026-04-21",
        "2026-04-22",
        "2026-04-23",
    ]
    assert str(rows[0][2]) == "消防署 / 災害管理組"
    assert str(rows[2][2]) == "消防署 / 災害管理組"
    assert rows[3][2] is None


def test_export_splits_fullwidth_department_separator(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 4, 20), date(2026, 4, 26)),
    )

    output_path = export_to_excel(
        [
            {
                "source": "金管會",
                "date": "2026-04-23",
                "department": "金管會／證券期貨局證券發行組",
                "title": "金管會新聞",
                "link": "https://example.com/fsc",
            },
        ],
        output_dir=tmp_path,
    )

    workbook = load_workbook(output_path, rich_text=True)
    rows = list(workbook["全部新聞"].iter_rows(min_row=2, values_only=True))

    assert rows[0][0] == "金管會"
    assert rows[0][2] == "證券期貨局證券發行組"


def test_export_groups_digital_agencies_under_moda(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 5, 18), date(2026, 5, 24)),
    )

    output_path = export_to_excel(
        [
            {
                "source": "資通安全署",
                "date": "2026-05-21",
                "department": "資通安全署",
                "title": "資安署新聞",
                "link": "https://example.com/acs",
            },
            {
                "source": "數位產業署",
                "date": "2026-05-20",
                "department": "數位產業署",
                "title": "數產署新聞",
                "link": "https://example.com/adi",
            },
            {
                "source": "國家資通安全研究院",
                "date": "2026-05-22",
                "department": "國家資通安全研究院",
                "title": "資安院新聞",
                "link": "https://example.com/nics",
            },
        ],
        output_dir=tmp_path,
    )

    workbook = load_workbook(output_path, rich_text=True)
    rows = list(workbook["全部新聞"].iter_rows(min_row=2, values_only=True))

    assert [row[0] for row in rows] == ["數位發展部", "數位發展部", "數位發展部"]
    assert [str(row[2]) for row in rows] == ["數位產業署", "資通安全署", "國家資通安全研究院"]


def test_export_ignores_url_like_department_values(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 4, 20), date(2026, 4, 26)),
    )

    output_path = export_to_excel(
        [
            {
                "source": "法務部",
                "date": "2026-04-23",
                "department": "法務部／http://www.moj.gov.tw/2204/2795/2796/276188/post",
                "title": "法務部新聞",
                "link": "http://www.moj.gov.tw/2204/2795/2796/276188/post",
            },
        ],
        output_dir=tmp_path,
    )

    workbook = load_workbook(output_path, rich_text=True)
    rows = list(workbook["全部新聞"].iter_rows(min_row=2, values_only=True))

    assert rows[0][0] == "法務部"
    assert rows[0][2] is None


def test_export_dedupes_affiliated_cross_posts(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 4, 20), date(2026, 4, 26)),
    )

    title = "因應極端氣候複合式災害 劉世芳：精進防汛整備與強化救災韌性 確保國人生命財產安全"
    output_path = export_to_excel(
        [
            {
                "source": "內政部",
                "date": "2026-04-23",
                "department": "內政部／消防署",
                "title": title,
                "link": "https://www.moi.gov.tw/News_Content.aspx?n=4&s=338211",
            },
            {
                "source": "消防署",
                "date": "2026-04-23",
                "department": "消防署／整備應變組",
                "title": title,
                "link": "https://www.nfa.gov.tw/pro/index.php?code=list&flag=detail&ids=1470&article_id=21148",
            },
        ],
        output_dir=tmp_path,
        dedupe_affiliated=True,
    )

    workbook = load_workbook(output_path, rich_text=True)
    rows = list(workbook["全部新聞"].iter_rows(min_row=2, values_only=True))

    assert len(rows) == 1
    assert rows[0][0] == "內政部"
    assert str(rows[0][2]) == "消防署 / 整備應變組"
    assert str(rows[0][4]).startswith("消防署官網：https://www.nfa.gov.tw/")


def test_export_keeps_affiliated_cross_posts_by_default(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 4, 20), date(2026, 4, 26)),
    )

    title = "同一則消防新聞"
    output_path = export_to_excel(
        [
            {
                "source": "內政部",
                "date": "2026-04-23",
                "department": "內政部／消防署",
                "title": title,
                "link": "https://www.moi.gov.tw/news",
            },
            {
                "source": "消防署",
                "date": "2026-04-23",
                "department": "消防署／整備應變組",
                "title": title,
                "link": "https://www.nfa.gov.tw/news",
            },
        ],
        output_dir=tmp_path,
    )

    workbook = load_workbook(output_path, rich_text=True)
    rows = list(workbook["全部新聞"].iter_rows(min_row=2, values_only=True))

    assert len(rows) == 2


def test_export_adds_date_dropdowns_and_labeled_hyperlinks(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 6, 1), date(2026, 6, 7)),
    )

    output_path = export_to_excel(
        [
            {
                "source": "國科會",
                "date": "2026-06-04",
                "department": "國科會",
                "title": "國科會新聞",
                "link": "https://www.nstc.gov.tw/news",
            },
        ],
        output_dir=tmp_path,
    )

    workbook = load_workbook(output_path, rich_text=True)
    worksheet = workbook["全部新聞"]

    assert worksheet["B2"].value == "2026-06-04"
    assert str(worksheet["E2"].value) == "國科會官網：https://www.nstc.gov.tw/news"
    assert worksheet["E2"].hyperlink.target == "https://www.nstc.gov.tw/news"

    validations = list(worksheet.data_validations.dataValidation)
    assert any(
        "B2" in str(validation.sqref)
        and validation.formula1 == '"2026-06-04,民國115/6/4"'
        for validation in validations
    )


def test_export_styles_only_url_portion_of_labeled_hyperlinks(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 6, 1), date(2026, 6, 7)),
    )

    output_path = export_to_excel(
        [
            {
                "source": "國科會",
                "date": "2026-06-04",
                "department": "國科會",
                "title": "國科會新聞",
                "link": "https://www.nstc.gov.tw/news",
            },
        ],
        output_dir=tmp_path,
    )

    workbook = load_workbook(output_path, rich_text=True)
    cell = workbook["全部新聞"]["E2"]

    assert str(cell.value) == "國科會官網：https://www.nstc.gov.tw/news"
    assert cell.value.as_list() == ["國科會官網：", "https://www.nstc.gov.tw/news"]
    assert cell.hyperlink.target == "https://www.nstc.gov.tw/news"


def test_export_adds_ai_policy_metadata_and_two_relevance_colors(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 6, 1), date(2026, 6, 7)),
    )

    output_path = export_to_excel(
        [
            {
                "source": "國科會",
                "date": "2026-06-01",
                "department": "國科會",
                "title": "主權AI及算力建設正式啟動",
                "link": "https://example.com/high",
            },
            {
                "source": "國科會",
                "date": "2026-06-02",
                "department": "國科會",
                "title": "AI應用成果交流會",
                "link": "https://example.com/possible",
            },
            {
                "source": "國科會",
                "date": "2026-06-03",
                "department": "國科會",
                "title": "一般行政公告",
                "link": "https://example.com/unrelated",
            },
        ],
        output_dir=tmp_path,
    )

    workbook = load_workbook(output_path, rich_text=True)
    worksheet = workbook["全部新聞"]
    headers = [cell.value for cell in worksheet[1]]

    assert headers == [
        "部會", "新聞日期", "單位分類", "新聞標題", "新聞連結",
        "新聞摘要", "日期來源", "AI新十大建設", "主政部會", "關聯性", "關聯分數",
        "判定理由", "命中關鍵字", "排除關鍵字", "各建設評分",
    ]
    assert worksheet["G2"].value is None
    assert str(worksheet["H2"].value) == "主權AI及算力建設"
    assert worksheet["I2"].value == "國科會"
    assert worksheet["J2"].value == "高度相關"
    assert worksheet["K2"].value == 100
    assert "標題命中完整建設名稱" in str(worksheet["L2"].value)
    assert worksheet["J3"].value == "可能相關"
    assert worksheet["K3"].value == 40
    assert worksheet["J4"].value is None
    assert str(worksheet["O2"].value) == "主權AI及算力建設（100分，高度相關）"
    assert all(cell.fill.fgColor.rgb.endswith("FFFF00") for cell in worksheet[2])
    assert all(cell.fill.fgColor.rgb.endswith("FFF2CC") for cell in worksheet[3])
    assert all(cell.fill.fill_type is None for cell in worksheet[4])

    filtered_rows = list(workbook["已初步篩選工作表"].iter_rows(min_row=2, values_only=True))
    assert [row[9] for row in filtered_rows] == ["高度相關", "可能相關"]


def test_export_includes_ai_ten_initiatives_reference_sheet(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 6, 1), date(2026, 6, 7)),
    )

    output_path = export_to_excel([], output_dir=tmp_path)

    workbook = load_workbook(output_path, rich_text=True)
    rows = list(workbook["AI新十大建設對照"].iter_rows(min_row=2, values_only=True))
    names = [str(row[0]) for row in rows]
    lead_agencies = {str(row[0]): str(row[1]) for row in rows}

    assert len(names) == 10
    assert names[0] == "全民智慧生活圈"
    assert names[-1] == "千億資金驅動創新"
    assert lead_agencies["AI數位產業登峰"] == "數發部"
    assert lead_agencies["矽光子技術全球領先"] == "經濟部"


def test_export_uses_summary_for_possible_relevance(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "news_scraper.excel_exporter.get_cached_week_range",
        lambda: (date(2026, 6, 1), date(2026, 6, 7)),
    )

    output_path = export_to_excel(
        [
            {
                "source": "經濟部",
                "date": "2026-06-04",
                "department": "經濟部",
                "title": "前瞻技術計畫正式啟動",
                "link": "https://example.com/summary-match",
                "summary": "本計畫將推動矽光子技術全球領先並建置驗證場域。",
            },
        ],
        output_dir=tmp_path,
    )

    workbook = load_workbook(output_path, rich_text=True)
    row = workbook["全部新聞"][2]

    assert str(row[5].value) == "本計畫將推動矽光子技術全球領先並建置驗證場域。"
    assert str(row[7].value) == "矽光子技術全球領先"
    assert row[9].value == "可能相關"
    assert row[10].value == 70
    assert "摘要命中完整建設名稱" in str(row[11].value)
    assert str(row[14].value) == "矽光子技術全球領先（70分，可能相關）"
