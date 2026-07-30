import logging
import os
import re
import tempfile
from contextlib import contextmanager
from datetime import datetime
from importlib import import_module
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .color_utils import readable_text_color
from .config import AFFILIATED_SOURCE_PATHS, SOURCE_ORDER
from .relevance import (
    RelevanceProfile,
    build_default_relevance_profile,
    classify_relevance,
    get_relevance_profile_summary,
)
from .utils.dates import ad_date_to_roc_compact_str, ad_to_roc_str, get_cached_week_range
from .utils.dedupe import dedupe_affiliated_news
from .utils.text import clean_text, normalize_department_metadata_text

logger = logging.getLogger(__name__)
RELEVANCE_HIGH_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
RELEVANCE_POSSIBLE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
AI_POLICY_HIGHLIGHT_FILL = RELEVANCE_HIGH_FILL
AI_POLICY_POSSIBLE_FILL = RELEVANCE_POSSIBLE_FILL
HEADER_FONT = Font(name="標楷體", sz=11)
BODY_FONT = Font(name="Times New Roman", sz=11)
LINK_FONT = Font(name="Times New Roman", sz=11, color="0563C1", underline="single")
EAST_ASIAN_INLINE_FONT = InlineFont(rFont="標楷體", sz=11)
LATIN_INLINE_FONT = InlineFont(rFont="Times New Roman", sz=11)
LINK_INLINE_FONT = InlineFont(rFont="Times New Roman", sz=11, color="0563C1", u="single")
COLUMN_WIDTHS = {
    "A": 23.2,
    "B": 28,
    "C": 45,
    "D": 120,
    "E": 130,
    "F": 90,
    "G": 22,
    "H": 36,
    "I": 16,
    "J": 14,
    "K": 12,
    "L": 65,
    "M": 55,
    "N": 32,
    "O": 65,
}
ROW_HEIGHT = 22
BASE_COLUMNS = ["source", "date", "department", "title", "link", "summary", "date_source"]
EXPORT_COLUMNS = [
    "部會",
    "新聞日期",
    "單位分類",
    "新聞標題",
    "新聞連結",
    "新聞摘要",
    "日期來源",
    "關聯主題",
    "優先關聯機關",
    "關聯性",
    "關聯分數",
    "判定理由",
    "命中關鍵字",
    "排除關鍵字",
    "各主題評分",
]


@contextmanager
def atomic_excel_writer(output_path):
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=destination.parent,
        prefix=".{}-".format(destination.stem),
        suffix=".xlsx",
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook = Workbook()
        workbook.remove(workbook.active)
        yield workbook
        workbook.save(temporary_path)
        os.replace(temporary_path, destination)
    finally:
        temporary_path.unlink(missing_ok=True)


def split_department_path(department):
    normalized = normalize_department_metadata_text(department)
    if not normalized:
        return []
    parts = re.split(r"\s*／\s*|\s+/\s+", normalized)
    return [
        clean_part
        for part in parts
        if (clean_part := normalize_department_metadata_text(part))
    ]


def build_agency_path(source, department):
    source_name = clean_text(source)
    base_path = list(AFFILIATED_SOURCE_PATHS.get(source_name, (source_name,) if source_name else ()))
    department_path = split_department_path(department)

    path = list(base_path)
    for part in department_path:
        if part and part not in path:
            path.append(part)
    return path


def get_parent_source(source):
    path = build_agency_path(source, "")
    return path[0] if path else clean_text(source)


def format_department_path(source, department):
    path = build_agency_path(source, department)
    if len(path) <= 1:
        return ""
    return " / ".join(path[1:])


def agency_sort_key(row):
    source_name = clean_text(row.get("原始來源", ""))
    parent_source = clean_text(row.get("部會", ""))
    department_path = clean_text(row.get("單位分類", ""))
    return (
        SOURCE_ORDER.get(parent_source, SOURCE_ORDER.get(source_name, 999)),
        clean_text(row.get("新聞日期", "")),
        SOURCE_ORDER.get(source_name, 999),
        department_path,
        clean_text(row.get("新聞標題", "")),
    )


def format_news_link_display(source, link):
    source_name = clean_text(source)
    link_text = clean_text(link)
    if not link_text:
        return ""
    if not source_name:
        return link_text
    return "{}官網：{}".format(source_name, link_text)


def extract_news_link_url(link_text):
    text = clean_text(link_text)
    if not text:
        return ""
    url_match = re.search(r"https?://\S+", text)
    if url_match:
        return url_match.group(0)
    return text


def split_news_link_display_text(link_text):
    text = clean_text(link_text)
    if not text:
        return "", ""

    url_match = re.search(r"https?://\S+", text)
    if not url_match:
        return text, ""

    prefix = text[:url_match.start()]
    url = url_match.group(0)
    return prefix, url


def _coerce_records(value):
    if value is None:
        return []
    to_dict = getattr(value, "to_dict", None)
    if callable(to_dict):
        try:
            return [dict(row) for row in to_dict(orient="records")]
        except TypeError:
            pass
    return [dict(row) for row in value]


def _restore_optional_dataframe(original, rows):
    if type(original).__module__.split(".", 1)[0] != "pandas":
        return rows
    pandas = import_module("pandas")
    return pandas.DataFrame(rows)


def prepare_export_rows(rows):
    prepared = []
    for input_row in rows:
        row = dict(input_row)
        original_source = clean_text(row.get("原始來源") or row.get("部會"))
        row["原始來源"] = original_source
        row["部會"] = get_parent_source(original_source)
        row["單位分類"] = format_department_path(original_source, row.get("單位分類", ""))
        prepared.append(row)
    prepared.sort(key=agency_sort_key)
    for row in prepared:
        row.pop("原始來源", None)
    return prepared


def prepare_export_dataframe(df):
    """Compatibility wrapper for callers that still pass a pandas DataFrame."""

    rows = prepare_export_rows(_coerce_records(df))
    return _restore_optional_dataframe(df, rows)


def add_relevance_metadata_rows(rows, profile):
    enriched_rows = []
    topics_by_id = {topic.id: topic for topic in profile.topics}
    topic_order = {topic.id: index for index, topic in enumerate(profile.topics)}
    for input_row in rows:
        row = dict(input_row)
        result = classify_relevance(
            row.get("新聞標題", ""),
            source=row.get("部會", ""),
            summary=row.get("新聞摘要", ""),
            profile=profile,
        )
        dominant_match: dict[str, Any] = min(
            result["topic_matches"],
            key=lambda match: (
                -int(match.get("score") or 0),
                topic_order.get(str(match.get("id") or ""), 9999),
            ),
            default={},
        )
        dominant_topic = topics_by_id.get(str(dominant_match.get("id") or ""))
        row.update(
            {
                "關聯主題": "、".join(result["topics"]),
                "優先關聯機關": "、".join(result["priority_sources"]),
                "關聯性": result["relevance"],
                "關聯分數": result["score"],
                "判定理由": "；".join(result["reasons"]),
                "命中關鍵字": "、".join(result["matched_keywords"]),
                "排除關鍵字": "、".join(result["excluded_keywords"]),
                "各主題評分": "；".join(
                    "{}（{}分，{}）".format(match["name"], match["score"], match["relevance"])
                    for match in result["topic_matches"]
                ),
                "_主題識別色": dominant_topic.display_color if dominant_topic else "",
            }
        )
        enriched_rows.append(row)
    return enriched_rows


def add_relevance_metadata(df, profile):
    """Compatibility wrapper for callers that still pass a pandas DataFrame."""

    rows = add_relevance_metadata_rows(_coerce_records(df), profile)
    return _restore_optional_dataframe(df, rows)


def add_ai_policy_metadata(df):
    """Compatibility wrapper for callers that still use the v1.4 API."""

    return add_relevance_metadata(df, build_default_relevance_profile())


def build_relevance_reference_rows(profile):
    global_context = "、".join(
        rule.text for rule in profile.global_context_keywords if rule.enabled
    )
    global_exclusions = "、".join(
        rule.text
        for rule in profile.exclusions
        if rule.enabled and not rule.topic_id
    )
    return [
        {
            "主題": topic.name,
            "啟用": "是" if topic.enabled else "否",
            "優先關聯機關": "、".join(topic.priority_sources),
            "比對主題名稱": "是" if topic.match_name else "否",
            "顯示顏色": topic.display_color,
            "核心詞": "、".join(rule.text for rule in topic.core_keywords if rule.enabled),
            "輔助詞": "、".join(rule.text for rule in topic.supporting_keywords if rule.enabled),
            "主題脈絡詞": "、".join(rule.text for rule in topic.context_keywords if rule.enabled),
            "全域脈絡詞": global_context,
            "主題排除詞": "、".join(
                rule.text
                for rule in profile.exclusions
                if rule.enabled and rule.topic_id == topic.id
            ),
            "全域排除詞": global_exclusions,
        }
        for topic in profile.topics
    ]


def build_relevance_rules_rows(profile):
    rows = []
    for rule in profile.global_context_keywords:
        rows.append(
            {
                "主題": "全域",
                "規則類型": "脈絡詞",
                "關鍵字": rule.text,
                "比對欄位": "標題與摘要",
                "啟用": "是" if rule.enabled else "否",
                "來源": rule.origin,
                "規則ID": rule.id,
            }
        )
    for topic in profile.topics:
        for label, rules in (
            ("核心詞", topic.core_keywords),
            ("輔助詞", topic.supporting_keywords),
            ("脈絡詞", topic.context_keywords),
        ):
            for rule in rules:
                rows.append(
                    {
                        "主題": topic.name,
                        "規則類型": label,
                        "關鍵字": rule.text,
                        "比對欄位": "標題與摘要",
                        "啟用": "是" if rule.enabled else "否",
                        "來源": rule.origin,
                        "規則ID": rule.id,
                    }
                )
    topic_names = {topic.id: topic.name for topic in profile.topics}
    for rule in profile.exclusions:
        fields = {
            ("title",): "標題",
            ("summary",): "摘要",
            ("summary", "title"): "標題與摘要",
            ("title", "summary"): "標題與摘要",
        }.get(tuple(rule.match_fields), "、".join(rule.match_fields))
        rows.append(
            {
                "主題": topic_names.get(rule.topic_id, "全域"),
                "規則類型": "排除詞",
                "關鍵字": rule.text,
                "比對欄位": fields,
                "啟用": "是" if rule.enabled else "否",
                "來源": rule.origin,
                "規則ID": rule.id,
            }
        )
    return rows


def build_relevance_version_rows(profile, source_label):
    summary = get_relevance_profile_summary(profile, source_label=source_label)
    labels = {
        "schema_version": "設定格式版本",
        "name": "設定名稱",
        "template_version": "範本版本",
        "ruleset_hash": "有效規則雜湊",
        "source": "設定來源",
        "topic_count": "主題總數",
        "enabled_topic_count": "啟用主題數",
        "disabled_topic_count": "停用主題數",
        "keyword_count": "關鍵字總數",
        "enabled_keyword_count": "啟用關鍵字數",
        "disabled_keyword_count": "停用關鍵字數",
        "exclusion_count": "排除詞總數",
        "enabled_exclusion_count": "啟用排除詞數",
        "disabled_exclusion_count": "停用排除詞數",
        "custom_item_count": "自訂項目數",
        "deleted_default_count": "已刪除預設項目數",
    }
    rows = [{"項目": labels[key], "內容": value} for key, value in summary.items()]
    rows.append(
        {
            "項目": "匯出時間",
            "內容": datetime.now().astimezone().isoformat(timespec="seconds"),
        }
    )
    return rows


def _append_worksheet(workbook, title, columns, rows):
    worksheet = workbook.create_sheet(title)
    worksheet.append(list(columns))
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])
    return worksheet


def get_ai_policy_row_fill(relevance):
    if clean_text(relevance) == "高度相關":
        return RELEVANCE_HIGH_FILL
    if clean_text(relevance) == "可能相關":
        return RELEVANCE_POSSIBLE_FILL
    return None


def apply_topic_identity_colors(worksheet, rows, header_map):
    topic_col_idx = header_map.get("關聯主題")
    if not topic_col_idx:
        return
    for row_idx, row in enumerate(rows, 2):
        color = clean_text(row.get("_主題識別色", ""))
        if not re.fullmatch(r"#[0-9A-Fa-f]{6}", color):
            continue
        text_color, _ratio = readable_text_color(color)
        cell = worksheet.cell(row=row_idx, column=topic_col_idx)
        cell.fill = PatternFill(fill_type="solid", fgColor=color[1:].upper())
        cell.font = Font(name="標楷體", sz=11, color=text_color[1:])


def export_to_excel(
    news_items,
    output_dir,
    dedupe_affiliated=False,
    *,
    relevance_profile: RelevanceProfile | None = None,
    relevance_profile_source: str = "內建預設範本",
):
    relevance_profile = relevance_profile or build_default_relevance_profile()
    if news_items is None:
        news_items = []
    elif dedupe_affiliated:
        news_items = dedupe_affiliated_news(news_items)

    rename_map = {
        "source": "部會",
        "date": "新聞日期",
        "department": "單位分類",
        "title": "新聞標題",
        "link": "新聞連結",
        "summary": "新聞摘要",
        "date_source": "日期來源",
    }
    rows = []
    for item in news_items:
        normalized = {column: item.get(column, "") for column in BASE_COLUMNS}
        row = {rename_map[column]: normalized[column] for column in BASE_COLUMNS}
        row["原始來源"] = row["部會"]
        row["新聞連結"] = format_news_link_display(row["原始來源"], row["新聞連結"])
        rows.append(row)
    rows = prepare_export_rows(rows)
    for row in rows:
        row["新聞日期"] = clean_text(row.get("新聞日期", ""))
    rows = add_relevance_metadata_rows(rows, relevance_profile)

    start_of_week, end_of_week = get_cached_week_range()
    file_name = "本週新聞整理（{}至{}）.xlsx".format(
        ad_date_to_roc_compact_str(start_of_week),
        ad_date_to_roc_compact_str(end_of_week),
    )
    output_path = Path(output_dir) / file_name
    output_path.parent.mkdir(parents=True, exist_ok=True)

    topic_order = {
        topic.name: position for position, topic in enumerate(relevance_profile.topics)
    }
    highlighted_rows = [
        dict(row)
        for row in rows
        if row.get("關聯性") in ("高度相關", "可能相關")
    ]
    highlighted_rows.sort(
        key=lambda row: (
            {"高度相關": 0, "可能相關": 1}.get(clean_text(row.get("關聯性")), 2),
            -int(row.get("關聯分數") or 0),
            topic_order.get(clean_text(row.get("關聯主題", "")).split("、")[0], 9999),
        )
    )

    relevance_reference_rows = build_relevance_reference_rows(relevance_profile)
    relevance_rules_rows = build_relevance_rules_rows(relevance_profile)
    relevance_version_rows = build_relevance_version_rows(
        relevance_profile,
        relevance_profile_source,
    )

    source_sheet_map = {
        "財政部": "財政部",
        "國發會": "國發會",
        "國科會": "國科會",
        "數位發展部": "數發部",
        "經濟部": "經濟部",
    }

    with atomic_excel_writer(output_path) as workbook:
        _append_worksheet(workbook, "全部新聞", EXPORT_COLUMNS, rows)
        _append_worksheet(workbook, "已初步篩選工作表", EXPORT_COLUMNS, highlighted_rows)
        _append_worksheet(
            workbook,
            "主題規則對照",
            [
                "主題",
                "啟用",
                "優先關聯機關",
                "比對主題名稱",
                "顯示顏色",
                "核心詞",
                "輔助詞",
                "主題脈絡詞",
                "全域脈絡詞",
                "主題排除詞",
                "全域排除詞",
            ],
            relevance_reference_rows,
        )
        _append_worksheet(
            workbook,
            "關聯性規則",
            ["主題", "規則類型", "關鍵字", "比對欄位", "啟用", "來源", "規則ID"],
            relevance_rules_rows,
        )
        _append_worksheet(workbook, "規則版本", ["項目", "內容"], relevance_version_rows)

        for source_name, sheet_name in source_sheet_map.items():
            source_rows = [dict(row) for row in rows if row.get("部會") == source_name]
            _append_worksheet(workbook, sheet_name, EXPORT_COLUMNS, source_rows)

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            last_col_letter = get_column_letter(worksheet.max_column)
            worksheet.auto_filter.ref = "A1:{}{}".format(last_col_letter, worksheet.max_row)

            header_map = {
                cell.value: col_idx
                for col_idx, cell in enumerate(worksheet[1], start=1)
            }
            apply_date_dropdowns_to_sheet(worksheet)
            apply_hyperlinks_to_sheet(worksheet)

            relevance_col_idx = header_map.get("關聯性")
            if relevance_col_idx:
                for row_idx in range(2, worksheet.max_row + 1):
                    relevance = worksheet.cell(row=row_idx, column=relevance_col_idx).value
                    row_fill = get_ai_policy_row_fill(relevance)
                    if row_fill is not None:
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = row_fill

            format_excel_worksheet(worksheet, header_map)
            if worksheet.title == "全部新聞":
                apply_topic_identity_colors(worksheet, rows, header_map)
            elif worksheet.title == "已初步篩選工作表":
                apply_topic_identity_colors(worksheet, highlighted_rows, header_map)
            elif worksheet.title in source_sheet_map.values():
                source_name = next(
                    name for name, title in source_sheet_map.items() if title == worksheet.title
                )
                apply_topic_identity_colors(
                    worksheet,
                    [row for row in rows if row.get("部會") == source_name],
                    header_map,
                )
            if worksheet.title == "主題規則對照":
                for column_letter, width in {
                    "A": 34,
                    "B": 12,
                    "C": 28,
                    "D": 16,
                    "E": 16,
                    "F": 72,
                    "G": 72,
                    "H": 72,
                    "I": 72,
                    "J": 55,
                    "K": 55,
                }.items():
                    worksheet.column_dimensions[column_letter].width = width
            elif worksheet.title == "關聯性規則":
                for column_letter, width in {
                    "A": 34,
                    "B": 16,
                    "C": 55,
                    "D": 18,
                    "E": 10,
                    "F": 12,
                    "G": 42,
                }.items():
                    worksheet.column_dimensions[column_letter].width = width
            elif worksheet.title == "規則版本":
                worksheet.column_dimensions["A"].width = 28
                worksheet.column_dimensions["B"].width = 80

    logger.info("Excel 已輸出：%s", output_path)
    return output_path


def apply_date_dropdowns_to_sheet(ws):
    if ws is None or ws.max_row < 2 or ws.max_column < 1:
        return

    date_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if clean_text(str(cell_value)) == "新聞日期":
            date_col_idx = col_idx
            break

    if date_col_idx is None:
        return

    cells_by_options: dict[tuple[str, str], list[Cell]] = {}
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=date_col_idx)
        ad_date = clean_text(str(cell.value)) if cell.value is not None else ""
        if not ad_date:
            continue
        roc_date = "民國{}".format(ad_to_roc_str(ad_date))
        cells_by_options.setdefault((ad_date, roc_date), []).append(cell)

    for (ad_date, roc_date), cells in cells_by_options.items():
        validation = DataValidation(type="list", formula1='"{},{}"'.format(ad_date, roc_date), allow_blank=False)
        validation.error = "請選擇西元日期或民國日期。"
        validation.errorTitle = "日期格式不正確"
        validation.prompt = "可選擇西元紀年或民國紀年。"
        validation.promptTitle = "新聞日期格式"
        ws.add_data_validation(validation)
        for cell in cells:
            validation.add(cell)


def apply_hyperlinks_to_sheet(ws):
    if ws is None or ws.max_row < 2 or ws.max_column < 1:
        return

    link_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if clean_text(str(cell_value)) == "新聞連結":
            link_col_idx = col_idx
            break

    if link_col_idx is None:
        return

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=link_col_idx)
        link = clean_text(str(cell.value)) if cell.value is not None else ""
        if not link:
            continue
        cell.hyperlink = extract_news_link_url(link)


def is_east_asian_character(char):
    codepoint = ord(char)
    return (
        0x2E80 <= codepoint <= 0x2EFF
        or 0x2F00 <= codepoint <= 0x2FDF
        or 0x3000 <= codepoint <= 0x303F
        or 0x3040 <= codepoint <= 0x30FF
        or 0x3100 <= codepoint <= 0x312F
        or 0x31A0 <= codepoint <= 0x31BF
        or 0x3400 <= codepoint <= 0x4DBF
        or 0x4E00 <= codepoint <= 0x9FFF
        or 0xF900 <= codepoint <= 0xFAFF
        or 0xFE30 <= codepoint <= 0xFE6F
        or 0xFF00 <= codepoint <= 0xFFEF
    )


def classify_text_kind(text):
    has_east_asian = False
    has_latin = False
    for char in str(text):
        if not char or char.isspace():
            continue
        if is_east_asian_character(char):
            has_east_asian = True
        else:
            has_latin = True
        if has_east_asian and has_latin:
            return "mixed"
    if has_east_asian:
        return "east_asian"
    return "latin"


def build_mixed_font_rich_text(text):
    normalized_text = str(text)
    if not normalized_text:
        return normalized_text

    runs = []
    current_kind = None
    current_chars: list[str] = []

    def flush_run():
        if not current_chars:
            return
        inline_font = EAST_ASIAN_INLINE_FONT if current_kind == "east_asian" else LATIN_INLINE_FONT
        runs.append(TextBlock(inline_font, "".join(current_chars)))

    for char in normalized_text:
        char_kind = "east_asian" if is_east_asian_character(char) else "latin"
        if char.isspace() and current_kind is not None:
            char_kind = current_kind
        if current_kind is None:
            current_kind = char_kind
            current_chars.append(char)
            continue
        if char_kind != current_kind:
            flush_run()
            current_chars = [char]
            current_kind = char_kind
            continue
        current_chars.append(char)

    flush_run()
    if len(runs) <= 1:
        return normalized_text
    return CellRichText(runs)


def apply_cell_font(cell, is_header=False):
    if cell is None:
        return

    value = cell.value
    if value is None or value == "":
        cell.font = HEADER_FONT if is_header else BODY_FONT
        return

    if is_header:
        cell.font = HEADER_FONT
        return

    if cell.hyperlink:
        prefix, url = split_news_link_display_text(value)
        if prefix and url:
            cell.value = CellRichText(
                [
                    TextBlock(EAST_ASIAN_INLINE_FONT, prefix),
                    TextBlock(LINK_INLINE_FONT, url),
                ]
            )
            return
        cell.font = LINK_FONT
        return

    text_kind = classify_text_kind(value)
    if text_kind == "east_asian":
        cell.font = HEADER_FONT
        return
    if text_kind == "latin":
        cell.font = BODY_FONT
        return

    rich_text = build_mixed_font_rich_text(value)
    if isinstance(rich_text, CellRichText):
        cell.value = rich_text
        return
    cell.font = BODY_FONT


def format_excel_worksheet(worksheet, header_map=None):
    if worksheet is None or worksheet.max_column < 1:
        return

    max_column = worksheet.max_column
    if header_map is None:
        header_map = {
            clean_text(cell.value): col_idx
            for col_idx, cell in enumerate(worksheet[1], start=1)
        }

    for column_letter, column_width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = column_width

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.row_dimensions[1].height = ROW_HEIGHT
    for cell in worksheet[1]:
        cell.alignment = header_alignment
        apply_cell_font(cell, is_header=True)

    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = cell_alignment
            apply_cell_font(cell)
        worksheet.row_dimensions[row_idx].height = ROW_HEIGHT
