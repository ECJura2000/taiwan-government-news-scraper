from __future__ import annotations

import argparse
from datetime import date, timedelta
import json
from pathlib import Path
import tempfile
import time

from openpyxl import load_workbook

from news_scraper.excel_exporter import export_to_excel


def build_items(row_count: int) -> list[dict]:
    return [
        {
            "source": "行政院",
            "date": (date(2026, 7, 27) + timedelta(days=index % 7)).isoformat(),
            "department": "行政院／測試單位",
            "title": "第 {} 則政府政策新聞".format(index),
            "link": "https://example.test/news/{}".format(index),
            "summary": "用於驗證大量資料匯出、字型、連結與排序。",
            "date_source": "published",
        }
        for index in range(row_count)
    ]


def run_benchmark(row_count: int, output_dir: Path) -> dict:
    started_at = time.perf_counter()
    output_path = export_to_excel(build_items(row_count), output_dir=output_dir)
    duration_seconds = time.perf_counter() - started_at
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    exported_rows = workbook["全部新聞"].max_row - 1
    workbook.close()
    if exported_rows != row_count:
        raise ValueError("Excel 筆數不符：預期 {}，實際 {}".format(row_count, exported_rows))
    return {
        "row_count": row_count,
        "duration_seconds": round(duration_seconds, 3),
        "output_bytes": output_path.stat().st_size,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Benchmark pure-openpyxl Excel export.")
    parser.add_argument("--rows", type=int, default=10_000)
    parser.add_argument("--max-seconds", type=float, default=30)
    args = parser.parse_args(argv)
    with tempfile.TemporaryDirectory(prefix="news-scraper-excel-benchmark-") as temp_dir:
        result = run_benchmark(args.rows, Path(temp_dir))
    print(json.dumps(result, ensure_ascii=False))
    if result["duration_seconds"] > args.max_seconds:
        raise SystemExit(
            "Excel 匯出 {:.3f} 秒，超過 {:.1f} 秒門檻".format(
                result["duration_seconds"],
                args.max_seconds,
            )
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
